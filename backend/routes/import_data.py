"""
Data Import routes — Two modes:
  1. Opening Balance Entry  (POST /import/opening-balance)
  2. Excel Bulk Import      (POST /import/excel/preview  +  POST /import/excel/confirm)
  3. Template download      (GET  /import/template)
"""
import io, math
from datetime import datetime, timezone, date as date_type
from typing import Optional, List
from fastapi import APIRouter, HTTPException, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from bson import ObjectId
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from core.database import db
from core.auth import get_current_user
from helpers import _doc, generate_customer_id, generate_loan_number, _add_months

router = APIRouter()

# ─── Helpers ──────────────────────────────────────────────────────────────────

def _today_ym() -> str:
    t = date_type.today()
    return f"{t.year}-{t.month:02d}"


def _build_ob_schedule(opening_balance: float, emi_amount: Optional[float], start_ym: str) -> list:
    """Build EMI schedule for an opening-balance loan.
    All EMIs start from start_ym (current month).
    Returns empty list if emi_amount is None or <= 0 (Gyal loans).
    """
    if not emi_amount or emi_amount <= 0:
        return []
    n = math.ceil(opening_balance / emi_amount)
    schedule = []
    y, m = int(start_ym[:4]), int(start_ym[5:7])
    for i in range(n):
        due = _add_months(date_type(y, m, 1), i)
        amt = emi_amount if i < n - 1 else round(opening_balance - emi_amount * (n - 1), 2)
        schedule.append({
            "due_month": f"{due.year}-{due.month:02d}",
            "amount": amt,
            "status": "due",
            "paid_date": None,
        })
    return schedule


async def _resolve_illaka_misal(illaka_name: str, misal_name: str, misal_display_order: int = None):
    """Return (illaka_id, misal_id) resolving by name (case-insensitive).
    Creates the Misal if it doesn't exist in the Illaka.
    Raises ValueError if the Illaka is not found.
    """
    illaka = await db.illakas.find_one(
        {"name": {"$regex": f"^{illaka_name.strip()}$", "$options": "i"}}
    )
    if not illaka:
        raise ValueError(f"Illaka '{illaka_name}' not found")
    illaka_id = str(illaka["_id"])

    misal = await db.misals.find_one(
        {"illaka_id": illaka_id, "name": {"$regex": f"^{misal_name.strip()}$", "$options": "i"}}
    )
    if not misal:
        now = datetime.now(timezone.utc).isoformat()
        new_doc = {
            "name": misal_name.strip(),
            "illaka_id": illaka_id,
            "description": "",
            "created_at": now,
            "updated_at": now,
        }
        if misal_display_order is not None:
            new_doc["display_order"] = misal_display_order
        result = await db.misals.insert_one(new_doc)
        misal_id = str(result.inserted_id)
    else:
        misal_id = str(misal["_id"])

    return illaka_id, misal_id


async def _create_ob_kyc_and_loan(
    illaka_id: str,
    misal_id: str,
    client_name: str,
    client_phone: Optional[str],
    co_borrower_name: Optional[str],
    guarantor_name: Optional[str],
    loan_date: str,
    opening_balance: float,
    emi_amount: Optional[float],
    created_by_id: str,
    created_by_name: str,
    display_order: int = None,
) -> dict:
    """Create a minimal KYC + opening-balance loan. Returns loan doc."""
    now = datetime.now(timezone.utc).isoformat()
    illaka = await db.illakas.find_one({"_id": ObjectId(illaka_id)})
    illaka_name = illaka["name"] if illaka else illaka_id

    # ── KYC ─────────────────────────────────────────────────────────────────
    customer_id = await generate_customer_id(illaka_name)
    kyc_doc = {
        "kyc_number": customer_id,
        "customer_id": customer_id,
        "illaka_id": illaka_id,
        "misal_id": misal_id,
        "status": "approved",
        "is_import": True,
        "primary_borrower": {
            "name": client_name.strip(),
            "phone": (client_phone or "").strip(),
            "address": "",
            "father_husband_name": "",
        },
        "co_borrower": {
            "name": co_borrower_name.strip(), "phone": "", "address": "", "father_husband_name": ""
        } if co_borrower_name and co_borrower_name.strip() else None,
        "guarantor": {
            "name": guarantor_name.strip(), "phone": "", "address": "", "father_husband_name": ""
        } if guarantor_name and guarantor_name.strip() else None,
        "created_at": now,
        "updated_at": now,
        "created_by_id": created_by_id,
        "created_by_name": created_by_name,
    }

    kyc_result = await db.kycs.insert_one(kyc_doc)
    kyc_id = str(kyc_result.inserted_id)

    # ── Loan ────────────────────────────────────────────────────────────────
    loan_number = await generate_loan_number(customer_id, kyc_id)
    start_ym = _today_ym()
    is_gyal = not emi_amount or emi_amount <= 0
    schedule = _build_ob_schedule(opening_balance, emi_amount, start_ym)

    loan_doc = {
        "loan_number": loan_number,
        "customer_id": customer_id,
        "kyc_id": kyc_id,
        "illaka_id": illaka_id,
        "misal_id": misal_id,
        "client_name": client_name.strip(),
        "client_phone": (client_phone or "").strip(),
        "loan_date": loan_date,
        "loan_type": "opening_balance",
        "principal_amount": opening_balance,
        "total_repayable": opening_balance,
        "emi_amount": emi_amount or 0,
        "emi_count": len(schedule),
        "emi_schedule": schedule,
        "status": "active",
        "is_gyal": is_gyal,
        "gyal_since": _today_ym() if is_gyal else None,
        "is_import": True,
        "total_paid": 0.0,
        "created_at": now,
        "updated_at": now,
        "created_by_id": created_by_id,
        "created_by_name": created_by_name,
    }
    if display_order is not None:
        loan_doc["display_order"] = display_order
    await db.loans.insert_one(loan_doc)
    return {"loan_number": loan_number, "kyc_id": kyc_id, "emi_count": len(schedule), "is_gyal": is_gyal}


# ─── Models ───────────────────────────────────────────────────────────────────

class OpeningBalanceEntry(BaseModel):
    illaka_id: str
    misal_id: str
    client_name: str
    client_phone: Optional[str] = None
    co_borrower_name: Optional[str] = None
    guarantor_name: Optional[str] = None
    loan_date: str           # YYYY-MM-DD
    opening_balance: float
    emi_amount: Optional[float] = None


class ExcelRow(BaseModel):
    illaka_name: str
    misal_name: str
    client_name: str
    client_phone: Optional[str] = None
    co_borrower_name: Optional[str] = None
    guarantor_name: Optional[str] = None
    loan_date: str
    opening_balance: float
    emi_amount: Optional[float] = None


class ExcelConfirmRequest(BaseModel):
    rows: List[ExcelRow]


# ─── Endpoints ────────────────────────────────────────────────────────────────

@router.get("/import/template")
async def download_template(request: Request):
    """Download Excel import template with instructions and a sample row."""
    await get_current_user(request)  # auth check

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Template"

    # Styles
    hdr_fill = PatternFill("solid", fgColor="1a1a2e")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    req_fill = PatternFill("solid", fgColor="fff3cd")
    opt_fill = PatternFill("solid", fgColor="d1ecf1")
    ex_fill  = PatternFill("solid", fgColor="f8f9fa")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center = Alignment(horizontal="center", vertical="center")

    headers = [
        ("Illaka Name *", True),
        ("Misal Name *", True),
        ("Client Name *", True),
        ("Client Phone", False),
        ("Co-borrower Name", False),
        ("Guarantor Name", False),
        ("Loan Date * (YYYY-MM-DD)", True),
        ("Opening Balance *", True),
        ("EMI Amount (blank = Gyal)", False),
    ]

    # Header row
    for col_idx, (label, required) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = thin
        cell.alignment = center

    # Legend row
    ws.merge_cells("A2:C2")
    ws["A2"] = "* Required fields (yellow). Optional fields are in light blue. EMI Amount blank = Gyal."
    ws["A2"].font = Font(italic=True, size=9, color="555555")
    ws["A2"].alignment = Alignment(horizontal="left")

    # Colour the columns to show required vs optional
    req_cols = [1, 2, 3, 7, 8]
    opt_cols = [4, 5, 6, 9]
    for col in req_cols:
        ws.cell(row=2, column=col).fill = req_fill
    for col in opt_cols:
        ws.cell(row=2, column=col).fill = opt_fill

    # Sample row (row 3)
    sample = [
        "Delhi", "Ward 1", "Ram Lal", "9876543210",
        "Shyam Lal", "Mohan Lal", "2024-06-15", 12000, 1000,
    ]
    for col_idx, val in enumerate(sample, start=1):
        cell = ws.cell(row=3, column=col_idx, value=val)
        cell.fill = ex_fill
        cell.border = thin

    # Column widths
    widths = [18, 18, 20, 16, 20, 20, 26, 18, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ("Bahi Khata — Data Import Instructions", True),
        ("", False),
        ("REQUIRED FIELDS", True),
        ("Illaka Name — must exactly match an existing Illaka in the system", False),
        ("Misal Name — must match an existing Misal; if not found it will be auto-created inside the Illaka", False),
        ("Client Name — full name of the primary borrower", False),
        ("Loan Date — format: YYYY-MM-DD  (e.g. 2024-06-15)", False),
        ("Opening Balance — outstanding amount remaining to be collected (₹)", False),
        ("", False),
        ("OPTIONAL FIELDS", True),
        ("Client Phone — 10-digit mobile number", False),
        ("Co-borrower Name — name of co-borrower (if any)", False),
        ("Guarantor Name — name of guarantor (if any)", False),
        ("EMI Amount — monthly instalment (₹). Leave BLANK to import as Gyal (bad debt) — no EMI schedule will be created.", False),
        ("", False),
        ("NOTES", True),
        ("• Do NOT modify column headers in the Import Template sheet.", False),
        ("• Start entering data from row 3 (replace the sample row or add below it).", False),
        ("• EMI schedule will start from the current month.", False),
        ("• Imported loans are marked as 'Opening Balance' type.", False),
        ("• If EMI Amount is blank, the account is imported directly as Gyal (bad debt).", False),
        ("• You can correct errors and re-upload before confirming.", False),
    ]
    for row_idx, (text, bold) in enumerate(instructions, start=1):
        cell = ws2.cell(row=row_idx, column=1, value=text)
        if bold:
            cell.font = Font(bold=True, size=12 if row_idx == 1 else 10)
        else:
            cell.font = Font(size=10)
    ws2.column_dimensions["A"].width = 75

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bahi_khata_import_template.xlsx"},
    )


@router.post("/import/opening-balance")
async def create_opening_balance(data: OpeningBalanceEntry, request: Request):
    """Create a single KYC + opening-balance loan from a form submission."""
    current_user = await get_current_user(request)
    if current_user["role"] not in ["admin", "maalik"]:
        raise HTTPException(status_code=403, detail="Only Admin or Maalik can import data")
    if data.opening_balance <= 0:
        raise HTTPException(status_code=400, detail="Opening balance must be > 0")
    if data.emi_amount is not None and data.emi_amount <= 0:
        raise HTTPException(status_code=400, detail="EMI amount must be > 0 (or leave blank for Gyal)")
    try:
        date_type.fromisoformat(data.loan_date)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid loan date format. Use YYYY-MM-DD")

    result = await _create_ob_kyc_and_loan(
        illaka_id=data.illaka_id,
        misal_id=data.misal_id,
        client_name=data.client_name,
        client_phone=data.client_phone,
        co_borrower_name=data.co_borrower_name,
        guarantor_name=data.guarantor_name,
        loan_date=data.loan_date,
        opening_balance=data.opening_balance,
        emi_amount=data.emi_amount,
        created_by_id=current_user["id"],
        created_by_name=current_user["name"],
    )
    return {"success": True, **result}


@router.post("/import/excel/preview")
async def excel_preview(request: Request, file: UploadFile = File(...)):
    """Parse an uploaded Excel file and return a preview of rows + any errors."""
    current_user = await get_current_user(request)
    if current_user["role"] not in ["admin", "maalik"]:
        raise HTTPException(status_code=403, detail="Only Admin or Maalik can import data")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Excel file. Please use the provided template.")

    COLS = ["illaka_name", "misal_name", "client_name", "client_phone",
            "co_borrower_name", "guarantor_name", "loan_date", "opening_balance", "emi_amount"]

    valid_rows = []
    error_rows = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue  # skip blank rows

        vals = {COLS[i]: (str(row[i]).strip() if row[i] is not None else "") for i in range(len(COLS))}

        errors = []
        # Required field checks
        for req in ["illaka_name", "misal_name", "client_name", "loan_date", "opening_balance"]:
            if not vals.get(req):
                errors.append(f"'{req}' is required")

        # Numeric checks
        try:
            ob = float(vals.get("opening_balance") or 0)
            if ob <= 0:
                errors.append("Opening balance must be > 0")
        except Exception:
            errors.append("Opening balance must be a number")
            ob = 0

        # EMI amount — optional; blank means Gyal
        raw_emi = vals.get("emi_amount", "").strip()
        if raw_emi:
            try:
                emi = float(raw_emi)
                if emi <= 0:
                    errors.append("EMI amount must be > 0 (or leave blank to mark as Gyal)")
            except Exception:
                errors.append("EMI amount must be a number (or leave blank to mark as Gyal)")
                emi = None
        else:
            emi = None  # blank → Gyal

        # Date check
        loan_date = vals.get("loan_date", "")
        try:
            # Accept YYYY-MM-DD or DD/MM/YYYY or datetime objects
            if "/" in loan_date:
                parts = loan_date.split("/")
                loan_date = f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
            date_type.fromisoformat(loan_date[:10])
            loan_date = loan_date[:10]
        except Exception:
            errors.append(f"Loan date '{loan_date}' is invalid. Use YYYY-MM-DD")

        if errors:
            error_rows.append({"row": row_idx, "data": vals, "errors": errors})
        else:
            is_gyal = emi is None
            emi_count = math.ceil(ob / emi) if emi else 0
            valid_rows.append({
                "row": row_idx,
                "illaka_name": vals["illaka_name"],
                "misal_name": vals["misal_name"],
                "client_name": vals["client_name"],
                "client_phone": vals.get("client_phone") or "",
                "co_borrower_name": vals.get("co_borrower_name") or "",
                "guarantor_name": vals.get("guarantor_name") or "",
                "loan_date": loan_date,
                "opening_balance": ob,
                "emi_amount": emi,
                "emi_count": emi_count,
                "is_gyal": is_gyal,
            })

    return {
        "total_rows": len(valid_rows) + len(error_rows),
        "valid_count": len(valid_rows),
        "error_count": len(error_rows),
        "valid_rows": valid_rows,
        "error_rows": error_rows,
    }


@router.post("/import/excel/confirm")
async def excel_confirm(data: ExcelConfirmRequest, request: Request):
    """Commit all validated rows from the Excel preview."""
    current_user = await get_current_user(request)
    if current_user["role"] not in ["admin", "maalik"]:
        raise HTTPException(status_code=403, detail="Only Admin or Maalik can import data")

    imported = []
    failed = []

    # Pre-compute misal display_order: first appearance per (illaka, misal) pair
    misal_order_map: dict = {}  # (illaka_name_lower, misal_name_lower) -> display_order
    misal_counter = 0
    for row in data.rows:
        key = (row.illaka_name.strip().lower(), row.misal_name.strip().lower())
        if key not in misal_order_map:
            misal_order_map[key] = misal_counter
            misal_counter += 1

    for i, row in enumerate(data.rows):
        try:
            misal_key = (row.illaka_name.strip().lower(), row.misal_name.strip().lower())
            misal_do = misal_order_map[misal_key]
            illaka_id, misal_id = await _resolve_illaka_misal(row.illaka_name, row.misal_name, misal_display_order=misal_do)
            result = await _create_ob_kyc_and_loan(
                illaka_id=illaka_id,
                misal_id=misal_id,
                client_name=row.client_name,
                client_phone=row.client_phone,
                co_borrower_name=row.co_borrower_name,
                guarantor_name=row.guarantor_name,
                loan_date=row.loan_date,
                opening_balance=row.opening_balance,
                emi_amount=row.emi_amount,
                created_by_id=current_user["id"],
                created_by_name=current_user["name"],
                display_order=i,
            )
            imported.append({**result, "client_name": row.client_name})
        except Exception as e:
            failed.append({"client_name": row.client_name, "error": str(e)})

    return {
        "imported_count": len(imported),
        "failed_count": len(failed),
        "imported": imported,
        "failed": failed,
    }
